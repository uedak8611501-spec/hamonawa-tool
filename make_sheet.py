"""
A4印刷用 操業記録シート（Excel）を生成してBytesIOで返すモジュール
"""
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.page import PageMargins


def _bdr(l="thin", r="thin", t="thin", b="thin"):
    def s(x): return Side(style=x) if x else None
    return Border(left=s(l), right=s(r), top=s(t), bottom=s(b))

def _fill(hex_color):
    return PatternFill("solid", start_color=hex_color, end_color=hex_color)

def _fnt(bold=False, size=11, color="000000"):
    return Font(name="Arial", bold=bold, size=size, color=color)

def _aln(h="center", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)


NAVY   = "1F3864"
BLUE   = "BDD7EE"
LBLUE  = "DEEAF1"
GRAY   = "F2F2F2"
YELLOW = "FFFF99"
WHITE  = "FFFFFF"
GREEN  = "E2EFDA"
DGREEN = "375623"
DBLUE  = "2E74B5"


def build_sheet(total_hachi: int = 20) -> BytesIO:
    """
    操業記録シートを生成しBytesIOで返す。
    total_hachi: 鉢数（最大40まで対応）
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "操業記録シート"

    # ── A4印刷設定 ──────────────────────────────────────────
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.orientation = "portrait"
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToHeight = 1
    ws.page_setup.fitToWidth = 1
    ws.page_margins = PageMargins(
        left=0.7, right=0.7, top=0.9, bottom=0.9, header=0.3, footer=0.3
    )
    ws.print_options.horizontalCentered = True

    # 列幅設定
    for i, w in enumerate([1.5, 9, 6, 9, 6, 9, 6, 9, 6, 4], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # ── タイトル ────────────────────────────────────────────
    ws.row_dimensions[1].height = 34
    ws.merge_cells("A1:J1")
    c = ws["A1"]
    c.value = "  延縄（ハモ）操業記録シート"
    c.font = Font(name="Arial", bold=True, size=16, color=WHITE)
    c.fill = _fill(NAVY)
    c.alignment = _aln("left")
    c.border = _bdr("medium", "medium", "medium", "medium")

    ws.row_dimensions[2].height = 16
    ws.merge_cells("A2:J2")
    c = ws["A2"]
    c.value = "記入後、写真を撮ってアプリにアップロードしてください"
    c.font = _fnt(False, 9, "595959")
    c.fill = _fill(LBLUE)
    c.alignment = _aln("center")

    # ── 基本情報 ────────────────────────────────────────────
    ws.row_dimensions[3].height = 6
    ws.row_dimensions[4].height = 20
    ws.row_dimensions[5].height = 28

    for lbl_r, val_r, label, hint in [
        ("B4:C4", "B5:C5", "操業日", "　　年　　月　　日"),
        ("D4:E4", "D5:E5", "エサ", ""),
        ("F4:G4", "F5:G5", "投入開始", "　　：　　"),
        ("H4:I4", "H5:I5", "揚げ終了", "　　：　　"),
    ]:
        ws.merge_cells(lbl_r); ws.merge_cells(val_r)
        lc = lbl_r.split(":")[0]; vc = val_r.split(":")[0]
        c = ws[lc]; c.value = label
        c.font = _fnt(True, 9, WHITE); c.fill = _fill(NAVY)
        c.alignment = _aln("center"); c.border = _bdr("medium","thin","medium","thin")
        v = ws[vc]; v.value = hint
        v.fill = _fill(YELLOW); v.font = _fnt(False, 13)
        v.alignment = _aln("center"); v.border = _bdr("medium","thin","thin","medium")

    ws["J4"].value = "総鉢数"
    ws["J4"].font = _fnt(True, 9, WHITE); ws["J4"].fill = _fill(NAVY)
    ws["J4"].alignment = _aln("center"); ws["J4"].border = _bdr("thin","medium","medium","thin")
    ws["J5"].value = f"  {total_hachi} 鉢"
    ws["J5"].fill = _fill(YELLOW); ws["J5"].font = _fnt(False, 13)
    ws["J5"].alignment = _aln("center"); ws["J5"].border = _bdr("thin","medium","thin","medium")

    # ── CTD環境データ ────────────────────────────────────────
    ws.row_dimensions[6].height = 6
    ws.row_dimensions[7].height = 20
    ws.row_dimensions[8].height = 20
    ws.row_dimensions[9].height = 28

    ws.merge_cells("B7:J7")
    c = ws["B7"]; c.value = "  CTD 環境データ"
    c.font = _fnt(True, 10, WHITE); c.fill = _fill(DBLUE)
    c.alignment = _aln("left"); c.border = _bdr("medium","medium","medium","thin")

    for col1, col2, label in [
        ("B","C","表層水温(℃)"), ("D","E","底水温(℃)"),
        ("F","G","表層塩分(psu)"), ("H","I","底層塩分(psu)"), ("J","J","実測水深(m)")
    ]:
        if col1 != col2:
            ws.merge_cells(f"{col1}8:{col2}8")
            ws.merge_cells(f"{col1}9:{col2}9")
        lb = "medium" if col1 == "B" else "thin"
        rb = "medium" if col2 == "J" else "thin"
        c = ws[f"{col1}8"]; c.value = label
        c.font = _fnt(True, 8, NAVY); c.fill = _fill(BLUE)
        c.alignment = _aln("center"); c.border = _bdr(lb, rb, "thin", "thin")
        v = ws[f"{col1}9"]
        v.fill = _fill(YELLOW); v.font = _fnt(False, 13)
        v.alignment = _aln("center"); v.border = _bdr(lb, rb, "thin", "medium")

    # ── 釣果表 ──────────────────────────────────────────────
    ws.row_dimensions[10].height = 6
    ws.row_dimensions[11].height = 20

    ws.merge_cells("B11:J11")
    c = ws["B11"]; c.value = "  鉢ごとの釣果"
    c.font = _fnt(True, 10, WHITE); c.fill = _fill(DGREEN)
    c.alignment = _aln("left"); c.border = _bdr("medium","medium","medium","thin")

    ws.row_dimensions[12].height = 18
    for col, label in [("B","鉢番号"),("C","釣果（匹）"),("F","鉢番号"),("G","釣果（匹）")]:
        c = ws[f"{col}12"]; c.value = label
        c.font = _fnt(True, 9, WHITE); c.fill = _fill(DGREEN); c.alignment = _aln("center")
    ws.merge_cells("C12:E12"); ws.merge_cells("G12:J12")
    for col in ["B","C","D","E","F","G","H","I","J"]:
        lb = "medium" if col == "B" else "thin"
        rb = "medium" if col == "J" else "thin"
        ws[f"{col}12"].border = _bdr(lb, rb, "thin", "thin")

    rows_needed = (total_hachi + 1) // 2  # 2列レイアウト
    for i in range(1, rows_needed + 1):
        row = 12 + i
        ws.row_dimensions[row].height = 20
        left_no  = i
        right_no = i + rows_needed

        # 左列
        if left_no <= total_hachi:
            bg = GREEN if left_no % 2 == 0 else WHITE
            c = ws[f"B{row}"]; c.value = f"第 {left_no} 鉢"
            c.font = _fnt(True, 10); c.fill = _fill(bg)
            c.alignment = _aln("center"); c.border = _bdr("medium","thin","thin","thin")
            ws.merge_cells(f"C{row}:E{row}")
            v = ws[f"C{row}"]
            v.fill = _fill(YELLOW); v.font = _fnt(False, 12)
            v.alignment = _aln("center"); v.border = _bdr("thin","medium","thin","thin")

        # 右列
        if right_no <= total_hachi:
            bg = GREEN if right_no % 2 == 0 else WHITE
            c = ws[f"F{row}"]; c.value = f"第 {right_no} 鉢"
            c.font = _fnt(True, 10); c.fill = _fill(bg)
            c.alignment = _aln("center"); c.border = _bdr("medium","thin","thin","thin")
            ws.merge_cells(f"G{row}:J{row}")
            v = ws[f"G{row}"]
            v.fill = _fill(YELLOW); v.font = _fnt(False, 12)
            v.alignment = _aln("center"); v.border = _bdr("thin","medium","thin","thin")

    # ── 備考 ────────────────────────────────────────────────
    note_start = 12 + rows_needed + 1
    ws.row_dimensions[note_start].height = 6
    ws.row_dimensions[note_start + 1].height = 20
    ws.merge_cells(f"B{note_start+1}:J{note_start+1}")
    c = ws[f"B{note_start+1}"]; c.value = "  備考・特記事項"
    c.font = _fnt(True, 10, WHITE); c.fill = _fill("7F7F7F")
    c.alignment = _aln("left"); c.border = _bdr("medium","medium","medium","thin")

    for r in range(note_start + 2, note_start + 7):
        ws.row_dimensions[r].height = 18
        ws.merge_cells(f"B{r}:J{r}")
        c = ws[f"B{r}"]; c.fill = _fill(GRAY)
        c.border = _bdr("medium","medium","thin","medium" if r == note_start+6 else "thin")

    # ── フッター ────────────────────────────────────────────
    footer_row = note_start + 7
    ws.row_dimensions[footer_row].height = 14
    ws.merge_cells(f"A{footer_row}:J{footer_row}")
    c = ws[f"A{footer_row}"]
    c.value = "延縄操業データ管理ツール  |  帰港後に写真を撮影してアプリにアップロードしてください"
    c.font = Font(name="Arial", size=8, color="808080", italic=True)
    c.alignment = _aln("center")

    # BytesIOで返す
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
